from __future__ import annotations

from pathlib import Path
from typing import Dict, List

from src.document_engine.contracts.document import Document, DocumentSource

try:
    import openpyxl
except ImportError as exc:
    openpyxl = None
    _OPENPYXL_IMPORT_ERROR = exc


class XlsxDocumentLoader:
    def load(self, file_path: str) -> Document:
        if openpyxl is None:
            raise ImportError(
                "Spreadsheet ingestion requires the optional dependency 'openpyxl'. "
                "Install it with `pip install openpyxl` to enable XLSX document support."
            ) from _OPENPYXL_IMPORT_ERROR

        path = Path(file_path)
        workbook = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        sheets: Dict[str, List[List[str]]] = {}
        for sheet_name in workbook.sheetnames[:3]:
            sheet = workbook[sheet_name]
            rows = []
            for row in sheet.iter_rows(values_only=True):
                rows.append([str(value).strip() if value is not None else "" for value in row])
                if len(rows) >= 100:
                    break
            sheets[sheet_name] = rows

        content_parts = ["\n".join(["\t".join(row) for row in rows]) for rows in sheets.values()]
        content = "\n\n".join(content_parts).strip()

        metadata: Dict[str, object] = {
            "source_type": "spreadsheet",
            "sheet_names": workbook.sheetnames,
            "sheet_count": len(workbook.sheetnames),
        }

        source = DocumentSource(
            path=str(path),
            source_type="spreadsheet",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        return Document(source=source, content=content, metadata=metadata)
