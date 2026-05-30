from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


@dataclass(slots=True)
class TableCell:
    value: Optional[str] = None
    row_index: int = 0
    col_index: int = 0
    is_header: bool = False

    def to_dict(self) -> Dict[str, Any]:
        return {
            "value": self.value,
            "row": self.row_index,
            "col": self.col_index,
            "is_header": self.is_header,
        }


@dataclass(slots=True)
class NormalizedTable:
    columns: List[str] = field(default_factory=list)
    rows: List[List[Optional[str]]] = field(default_factory=list)
    row_count: int = 0
    column_count: int = 0
    confidence_score: float = 0.0
    source_type: str = "unknown"
    quality_flags: List[str] = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "columns": self.columns,
            "rows": self.rows,
            "row_count": self.row_count,
            "column_count": self.column_count,
            "confidence_score": round(float(self.confidence_score), 2),
            "source_type": self.source_type,
            "quality_flags": self.quality_flags,
        }


class TableParser:
    def __init__(self):
        self.csv_delimiter_patterns = {",": 0, ";": 0, "\t": 0, "|": 0}

    def parse_csv_string(self, content: str) -> NormalizedTable:
        lines = content.split("\n")
        if not lines:
            return NormalizedTable(source_type="csv", confidence_score=0.0, quality_flags=["empty_input"])

        rows_list: List[List[str]] = []
        columns: List[str] = []
        delimiter = ","

        try:
            sample = "\n".join(lines[:5])
            dialect = csv.Sniffer().sniff(sample) if sample else csv.excel
            delimiter = dialect.delimiter
        except Exception:
            delimiter = ","

        quality_flags: List[str] = []
        row_count = 0

        try:
            reader = csv.reader(lines, delimiter=delimiter)
            for idx, row in enumerate(reader):
                cleaned_row = [cell.strip() for cell in row]
                if idx == 0:
                    columns = cleaned_row
                else:
                    rows_list.append(cleaned_row)
                row_count = idx + 1
                if row_count >= 100:
                    break
        except Exception as exc:
            quality_flags.append(f"csv_parse_error: {str(exc)}")

        confidence = 0.85 if not quality_flags else 0.6
        return NormalizedTable(
            columns=columns,
            rows=rows_list,
            row_count=len(rows_list),
            column_count=len(columns),
            confidence_score=confidence,
            source_type="csv",
            quality_flags=quality_flags,
        )

    def parse_xlsx_sheet(self, file_path: str, sheet_name: Optional[str] = None) -> NormalizedTable:
        if load_workbook is None:
            return NormalizedTable(
                source_type="xlsx",
                confidence_score=0.0,
                quality_flags=["openpyxl_not_installed"],
            )

        quality_flags: List[str] = []
        columns: List[str] = []
        rows_list: List[List[Optional[str]]] = []

        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            active_sheet = workbook[sheet_name] if sheet_name else workbook.active

            row_count = 0
            for row_idx, row in enumerate(active_sheet.iter_rows(values_only=True)):
                if row_idx == 0:
                    columns = [str(cell).strip() if cell else "" for cell in row]
                else:
                    cleaned_row = [str(cell).strip() if cell is not None else "" for cell in row]
                    rows_list.append(cleaned_row)
                row_count = row_idx + 1
                if row_count >= 100:
                    break
            workbook.close()
        except Exception as exc:
            quality_flags.append(f"xlsx_parse_error: {str(exc)}")

        confidence = 0.88 if not quality_flags else 0.6
        return NormalizedTable(
            columns=columns,
            rows=rows_list,
            row_count=len(rows_list),
            column_count=len(columns),
            confidence_score=confidence,
            source_type="xlsx",
            quality_flags=quality_flags,
        )

    def extract_text_table(self, content: str, min_columns: int = 2) -> List[NormalizedTable]:
        lines = [line.rstrip() for line in content.split("\n")]
        tables: List[NormalizedTable] = []
        current_table_lines: List[str] = []

        for line in lines:
            stripped = line.strip()

            if not stripped:
                if current_table_lines:
                    table = self._parse_text_table(current_table_lines, min_columns)
                    if table and table.column_count >= min_columns:
                        tables.append(table)
                    current_table_lines = []
            else:
                col_count = len(re.split(r"\s{2,}|\t|\|", stripped))
                if col_count >= min_columns:
                    current_table_lines.append(line)
                elif current_table_lines:
                    table = self._parse_text_table(current_table_lines, min_columns)
                    if table and table.column_count >= min_columns:
                        tables.append(table)
                    current_table_lines = []

        if current_table_lines:
            table = self._parse_text_table(current_table_lines, min_columns)
            if table and table.column_count >= min_columns:
                tables.append(table)

        return tables

    def _parse_text_table(self, lines: List[str], min_columns: int) -> NormalizedTable:
        if not lines:
            return NormalizedTable(
                confidence_score=0.0,
                quality_flags=["empty_table"],
                source_type="text",
            )

        columns: List[str] = []
        rows_list: List[List[str]] = []

        delimiter = self._detect_delimiter(lines[0])
        quality_flags: List[str] = []

        for idx, line in enumerate(lines):
            if delimiter in line:
                parts = [p.strip() for p in line.split(delimiter)]
            else:
                parts = re.split(r"\s{2,}|\t", line.strip())

            if idx == 0:
                columns = parts
            else:
                padded = parts + [""] * (len(columns) - len(parts)) if len(parts) < len(columns) else parts[: len(columns)]
                rows_list.append(padded)

        if len(columns) < min_columns:
            quality_flags.append(f"insufficient_columns: {len(columns)} < {min_columns}")

        confidence = 0.7 if not quality_flags else 0.4
        return NormalizedTable(
            columns=columns,
            rows=rows_list,
            row_count=len(rows_list),
            column_count=len(columns),
            confidence_score=confidence,
            source_type="text",
            quality_flags=quality_flags,
        )

    def _detect_delimiter(self, line: str) -> str:
        delimiters = [",", ";", "\t", "|"]
        for delim in delimiters:
            if line.count(delim) >= 2:
                return delim
        return ","
